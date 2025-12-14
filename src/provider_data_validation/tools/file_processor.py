"""
File processing utilities for extracting provider data from PDF and Excel files.
Supports extraction of provider names and related information.
"""

import io
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    import pypdf
except ImportError:
    pypdf = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


class FileProcessor:
    """Processes PDF and Excel files to extract provider information."""
    
    COMMON_HEADERS = [
        "name", "provider", "provider_name", "doctor", "physician",
        "full_name", "last_name", "first_name",
        "phone", "contact", "phone_number", "contact_number",
        "address", "location", "clinic", "facility",
        "license", "license_no", "license_number", "npi", "npi_number",
        "specialty", "specialization", "credentials",
        "hospital", "affiliation", "department"
    ]
    
    @staticmethod
    def _is_scanned_pdf(reader) -> bool:
        """Check if PDF is scanned/image-based (minimal extractable text)."""
        try:
            total_text = ""
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    total_text += text
            return len(total_text.strip()) < 50
        except:
            return True
    
    @staticmethod
    def _extract_with_ocr(file_content: bytes) -> str:
        """Extract text using vision LLM (preferred) or Tesseract OCR."""
        try:
            from pdf2image import convert_from_bytes
            
            print("📷 Converting PDF to images for OCR...")
            images = convert_from_bytes(file_content, dpi=300)
            full_text = ""
            
            # Try vision LLM first
            try:
                from .ocr_agent import extract_text_with_vision_llm, is_ollama_available
                
                if is_ollama_available():
                    print("🤖 Using LLaVA vision model...")
                    for i, image in enumerate(images):
                        print(f"  Page {i+1}/{len(images)}...")
                        text = extract_text_with_vision_llm(image)
                        full_text += text + "\n\n"
                    print(f"✓ Vision LLM complete ({len(full_text)} chars)")
                    return full_text
                else:
                    print("⚠️ LLaVA not available, using Tesseract...")
            except:
                print("⚠️ Vision LLM failed, using Tesseract...")
            
            # Fallback to Tesseract
            import pytesseract
            import platform
            if platform.system() == "Windows":
                paths = [r'C:\Program Files\Tesseract-OCR\tesseract.exe',
                        r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe']
                for path in paths:
                    if Path(path).exists():
                        pytesseract.pytesseract.tesseract_cmd = path
                        break
            
            for i, image in enumerate(images):
                print(f"  Page {i+1}/{len(images)}...")
                text = pytesseract.image_to_string(image, lang='eng')
                full_text += text + "\n\n"
            
            print(f"✓ OCR complete ({len(full_text)} chars)")
            return full_text
        except ImportError:
            raise ImportError(
                "OCR libraries not installed.\n"
                "Run: pip install pytesseract pdf2image Pillow\n"
                "Install Tesseract: https://github.com/UB-Mannheim/tesseract/wiki"
            )
        except Exception as e:
            raise ValueError(f"OCR failed: {str(e)}")
    
    @staticmethod
    def extract_from_pdf(file_content: bytes) -> List[Dict[str, Any]]:
        """
        Extract provider data from PDF file.
        Returns list of provider dictionaries.
        """
        if not pypdf:
            raise ImportError("pypdf is not installed. Install it with: pip install pypdf")
        
        providers = []
        
        try:
            pdf_file = io.BytesIO(file_content)
            reader = pypdf.PdfReader(pdf_file)
            
            # Check if PDF is scanned/image-based
            is_scanned = FileProcessor._is_scanned_pdf(reader)
            
            if is_scanned:
                print("📷 Scanned/handwritten PDF detected. Using OCR...")
                full_text = FileProcessor._extract_with_ocr(file_content)
            else:
                print("📄 Text-based PDF. Using standard extraction...")
                full_text = ""
                for page in reader.pages:
                    full_text += page.extract_text() + "\n"
            
            # Parse text to find provider entries
            print(f"\n{'='*60}")
            print(f"EXTRACTED PDF TEXT ({len(full_text)} chars):")
            print(f"{'='*60}")
            print(full_text[:500] if len(full_text) > 500 else full_text)  # First 500 chars
            print(f"{'='*60}\n")
            
            providers = FileProcessor._parse_text_for_providers(full_text)
            print(f"Found {len(providers)} providers")
            
        except Exception as e:
            print(f"Error extracting from PDF: {e}")
            raise ValueError(f"Failed to process PDF: {str(e)}")
        
        return providers
    
    @staticmethod
    def extract_from_excel(file_content: bytes) -> List[Dict[str, Any]]:
        """
        Extract provider data from Excel file.
        Supports .xlsx files.
        Returns list of provider dictionaries.
        """
        if not openpyxl:
            raise ImportError("openpyxl is not installed. Install it with: pip install openpyxl")
        
        providers = []
        
        try:
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file)
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Find header row (first row with data)
                header_row = None
                header_indices = {}
                
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if row and any(row):  # Skip empty rows
                        # Try to match headers
                        matched_headers = 0
                        for col_idx, cell_value in enumerate(row):
                            if cell_value:
                                cell_str = str(cell_value).lower().strip()
                                for header in FileProcessor.COMMON_HEADERS:
                                    if header in cell_str:
                                        header_indices[header] = col_idx
                                        matched_headers += 1
                        
                        if matched_headers >= 2:  # Found headers
                            header_row = row_idx
                            break
                
                # Extract data rows
                if header_row:
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        if row_idx <= header_row:
                            continue
                        
                        # Skip empty rows
                        if not row or not any(row):
                            continue
                        
                        provider_dict = {}
                        
                        # Extract provider info based on header positions
                        for header, col_idx in header_indices.items():
                            if col_idx < len(row):
                                cell_value = row[col_idx]
                                if cell_value:
                                    provider_dict[header] = str(cell_value).strip()
                        
                        # Standardize field names
                        provider = FileProcessor._standardize_provider_dict(provider_dict)
                        
                        # Only add if we have a provider name
                        if provider.get("provider_name"):
                            providers.append(provider)
                else:
                    # No headers found - assume first column is provider names
                    print("No headers found. Treating first column as provider names.")
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        # Skip empty rows
                        if not row or not any(row):
                            continue
                        
                        # Get first non-empty cell in row
                        provider_name = None
                        for cell_value in row:
                            if cell_value:
                                provider_name = str(cell_value).strip()
                                break
                        
                        if provider_name:
                            providers.append({"provider_name": provider_name})
            
        except Exception as e:
            print(f"Error extracting from Excel: {e}")
            raise ValueError(f"Failed to process Excel file: {str(e)}")
        
        return providers
    
    @staticmethod
    def _parse_text_for_providers(text: str) -> List[Dict[str, Any]]:
        """
        Parse text content to find provider entries.
        This is a basic implementation - enhance based on your text structure.
        """
        providers = []
        
        # Split by lines and look for patterns
        lines = text.split('\n')
        
        current_provider = {}
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Clean list markers (-, *, •, 1., etc.) before processing
            import re
            line = re.sub(r'^[-*•]\s*', '', line)  # Remove bullet points
            line = re.sub(r'^\d+\.?\s*', '', line)  # Remove numbers
            line = line.strip()
            
            if not line:
                continue
            
            # Check for common patterns
            line_lower = line.lower()
            
            # Look for name patterns (Dr., MD, etc.)
            if any(prefix in line for prefix in ['dr.', 'dr ', 'prof.', 'prof ']):
                # This might be a provider name
                if current_provider.get("provider_name"):
                    # Save previous provider
                    if current_provider:
                        providers.append(FileProcessor._standardize_provider_dict(current_provider))
                    current_provider = {}
                
                current_provider["provider_name"] = line.replace("Dr.", "Dr. ").replace("Prof.", "Prof. ")
            # Also check for simple names (2-3 words, capitalized, no numbers)
            elif (len(line.split()) >= 2 and len(line.split()) <= 4 and 
                  line[0].isupper() and not any(char.isdigit() for char in line) and
                  5 < len(line) < 50):
                if current_provider.get("provider_name"):
                    providers.append(FileProcessor._standardize_provider_dict(current_provider))
                    current_provider = {}
                current_provider["provider_name"] = line
            
            # Look for phone patterns
            elif any(indicator in line_lower for indicator in ['phone', 'tel:', 'contact:']):
                # Extract phone number
                import re
                phone_match = re.search(r'\d{3}[-.\s]?\d{3}[-.\s]?\d{4}', line)
                if phone_match:
                    current_provider["phone"] = phone_match.group()
            
            # Look for license/NPI patterns
            elif any(indicator in line_lower for indicator in ['license', 'npi', 'npi:']):
                import re
                number_match = re.search(r':\s*(\d+)', line)
                if number_match:
                    if 'npi' in line_lower:
                        current_provider["npi_number"] = number_match.group(1)
                    else:
                        current_provider["license_no"] = number_match.group(1)
        
        # Add last provider
        if current_provider.get("provider_name"):
            providers.append(FileProcessor._standardize_provider_dict(current_provider))
        
        return providers
    
    @staticmethod
    def _standardize_provider_dict(provider: Dict[str, Any]) -> Dict[str, Any]:
        """
        Standardize provider dictionary field names to match ProviderInput schema.
        """
        standardized = {}
        
        # Map common variations to standard field names
        field_mapping = {
            "provider_name": ["name", "provider", "provider_name", "doctor", "physician", "full_name"],
            "phone": ["phone", "contact", "phone_number", "contact_number", "telephone"],
            "address": ["address", "location", "clinic", "facility", "clinic_address"],
            "specialty": ["specialty", "specialization", "credentials", "medical_specialty"],
            "license_no": ["license", "license_no", "license_number", "license_num"],
            "npi_number": ["npi", "npi_number"],
            "hospital_affiliation": ["hospital", "affiliation", "hospital_affiliation", "department"]
        }
        
        for target_field, source_fields in field_mapping.items():
            for source_field in source_fields:
                if source_field in provider:
                    standardized[target_field] = provider[source_field]
                    break
        
        return standardized
    
    @staticmethod
    def get_file_type(filename: str) -> str:
        """Determine file type from filename."""
        filename_lower = filename.lower()
        
        if filename_lower.endswith('.pdf'):
            return 'pdf'
        elif filename_lower.endswith('.xlsx'):
            return 'xlsx'
        elif filename_lower.endswith('.xls'):
            return 'xls'
        else:
            return 'unknown'
    
    @staticmethod
    def validate_file(file_content: bytes, file_type: str) -> bool:
        """Validate file format and content."""
        if file_type == 'pdf':
            # Check PDF magic number
            return file_content.startswith(b'%PDF')
        elif file_type in ['xlsx', 'xls']:
            # Check ZIP magic number (XLSX is ZIP)
            return file_content.startswith(b'PK')
        return False
