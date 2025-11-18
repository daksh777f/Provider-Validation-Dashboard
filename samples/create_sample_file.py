"""
Generate a sample Excel file with provider names for testing.
"""
import openpyxl
from pathlib import Path

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Providers"

# Add headers
ws['A1'] = 'Provider Name'
ws['B1'] = 'Phone'
ws['C1'] = 'Specialty'

# Add sample providers that match the mock data
providers = [
    ['Dr Aarav Mehta', '8123456789', 'Cardiology'],
    ['Dr Priya Sharma', '9876543210', 'Dermatology'],
    ['Dr Rajesh Kumar', '7890123456', 'Neurology'],
]

# Add data rows
for idx, provider in enumerate(providers, start=2):
    ws[f'A{idx}'] = provider[0]
    ws[f'B{idx}'] = provider[1]
    ws[f'C{idx}'] = provider[2]

# Save the file
output_path = Path(__file__).parent / 'sample_providers.xlsx'
wb.save(output_path)
print(f"Sample file created: {output_path}")
